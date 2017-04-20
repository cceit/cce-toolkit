import csv
import datetime
import math
from functools import partial

import openpyxl
from openpyxl.utils import get_column_letter
import xlwt
from django.conf import settings
from django.core.files import temp as tempfile
from django.db.models import FieldDoesNotExist
from django.http import HttpResponse
from django.utils.text import capfirst


def get_width(string, bold=False):
    """
    Assuming a standard 10-point Arial font, returns the width of the string,
    in BIFF column width units.
    """
    char_widths = {
        '0': 262.637,
        '1': 262.637,
        '2': 262.637,
        '3': 262.637,
        '4': 262.637,
        '5': 262.637,
        '6': 262.637,
        '7': 262.637,
        '8': 262.637,
        '9': 262.637,
        'a': 262.637,
        'b': 262.637,
        'c': 262.637,
        'd': 262.637,
        'e': 262.637,
        'f': 146.015,
        'g': 262.637,
        'h': 262.637,
        'i': 117.096,
        'j': 88.178,
        'k': 233.244,
        'l': 88.178,
        'm': 379.259,
        'n': 262.637,
        'o': 262.637,
        'p': 262.637,
        'q': 262.637,
        'r': 175.407,
        's': 233.244,
        't': 117.096,
        'u': 262.637,
        'v': 203.852,
        'w': 321.422,
        'x': 203.852,
        'y': 262.637,
        'z': 233.244,
        'A': 321.422,
        'B': 321.422,
        'C': 350.341,
        'D': 350.341,
        'E': 321.422,
        'F': 291.556,
        'G': 350.341,
        'H': 321.422,
        'I': 146.015,
        'J': 262.637,
        'K': 321.422,
        'L': 262.637,
        'M': 379.259,
        'N': 321.422,
        'O': 350.341,
        'P': 321.422,
        'Q': 350.341,
        'R': 321.422,
        'S': 321.422,
        'T': 262.637,
        'U': 321.422,
        'V': 321.422,
        'W': 496.356,
        'X': 321.422,
        'Y': 321.422,
        'Z': 262.637,
        ' ': 146.015,
        '!': 146.015,
        '"': 175.407,
        '#': 262.637,
        '$': 262.637,
        '%': 438.044,
        '&': 321.422,
        '\'': 88.178,
        '(': 175.407,
        ')': 175.407,
        '*': 203.852,
        '+': 291.556,
        ',': 146.015,
        '-': 175.407,
        '.': 146.015,
        '/': 146.015,
        ':': 146.015,
        ';': 146.015,
        '<': 291.556,
        '=': 291.556,
        '>': 291.556,
        '?': 262.637,
        '@': 496.356,
        '[': 146.015,
        '\\': 146.015,
        ']': 146.015,
        '^': 203.852,
        '_': 262.637,
        '`': 175.407,
        '{': 175.407,
        '|': 146.015,
        '}': 175.407,
        '~': 291.556
    }

    length = 524  # Padding (Use multiples of 262)

    for char in string:
        if char in char_widths:
            length += char_widths[char]
        else:
            length += char_widths['0']

    if bold:
        length *= 1.1

    return int(math.ceil(length))  # Rounding up the decimal width value


def get_height(string):
    return int(len(string.split('\n')) * 220 * 1.2)


def is_datetime(obj):
    if type(obj) == datetime.date or type(obj) == datetime.datetime \
            or type(obj) == datetime:
        return True
    return False


def to_ascii(text):
    return text.encode('ascii', 'replace')


def write_to_worksheet(ws, row, column, cell):
    """
    Write a single cell to a worksheet with xlwt.
    Used with xls_multiple_worksheets_response.

    If "cell" is a dict and the key "merge" is present, the value of "merge"
    is also a dict with the potential to have keys called "row_span" and
    "col_span". These parameters indicate what cells (starting at row, column)
    should be merged together.

    :param cell: Simple or complex data to be written to the cell
    :type cell: str or dict with keys label, style and merge
    """
    if type(cell) is dict:
        label = cell.get('label')
        style = cell.get('style')
        merge = cell.get('merge')

        width = get_width(str(label))
        height = get_height(str(label))

        if merge is not None:
            row_span = merge.get('row_span', 0)
            col_span = merge.get('col_span', 0)
            width = 0
            if style is not None:
                if is_datetime(label):
                    ws.write_merge(row, row + row_span, column, column + col_span, label,
                                   xlwt.easyxf(num_format_str="YYYY-MM-DD"))
                else:
                    try:
                        ws.write_merge(row, row + row_span, column, column + col_span, label, xlwt.easyxf(style))
                    except:
                        ws.write_merge(row, row + row_span, column, column + col_span, str(label), xlwt.easyxf(style))

            else:
                if is_datetime(label):
                    ws.write_merge(row, row + row_span, column, column + col_span, label,
                                   xlwt.easyxf(num_format_str="YYYY-MM-DD"))
                try:
                    ws.write_merge(row, row + row_span, column, column + col_span, label)
                except:
                    ws.write_merge(row, row + row_span, column, column + col_span, str(label))
        else:
            if is_datetime(label):
                ws.write(row, column, label, xlwt.easyxf(style, num_format_str="YYYY-MM-DD"))
            if style is not None:
                try:
                    ws.write(row, column, label, xlwt.easyxf(style))
                except:
                    ws.write(row, column, str(label), xlwt.easyxf(style))
            else:
                try:
                    ws.write(row, column, label)
                except:
                    ws.write(row, column, str(label))

    elif isinstance(cell, tuple):
        label = cell[0]
        style = cell[1]
        width = get_width(str(label))
        height = get_height(str(label))

        if style is not None:
            if is_datetime(label):
                ws.write(row, column, label, xlwt.easyxf(style, num_format_str='YYYY-MM-DD'))
            else:
                try:
                    ws.write(row, column, label, xlwt.easyxf(style))
                except:
                    ws.write(row, column, str(label), xlwt.easyxf(style))
        else:
            try:
                ws.write(row, column, label)
            except:
                ws.write(row, column, str(label))

    else:
        label = cell
        width = get_width(str(label))
        height = get_height(str(label))
        if is_datetime(label):
            ws.write(row,
                     column,
                     label,
                     xlwt.easyxf(num_format_str='YYYY-MM-DD'))
        else:
            try:
                ws.write(row, column, label)
            except:
                label = str(label)
                ws.write(row, column, label)
    return width, height


def csv_response(filename, table):
    """Return a CSV file of the given table as an HttpResponse.

    Args:

        filename: the name of the downloaded CSV file. The extension will be
            '.csv'. This parameter is inserted directly to the response's
            Content-Disposition, and must be escaped accordingly.

        table: a 2-dimensional iterable, in row-major order.

    Returns:

        A CSV HttpResponse with appropriate content_type and
        Content-Disposition.

    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="%s.csv"' \
                                      % filename
    writer = csv.writer(response)
    for row in table:
        # Convert generators to lists for use by writer.writerow.
        writer.writerow(list(row))
    return response


def xls_response(filename, sheetname, table, header=None, footer=None,
                 include_totals=False, total_label='Total', grouper_col=None,
                 value_col=None):
    """Return a Microsoft Excel file of the given table as an HttpResponse.

    Args:

        filename: the name of the downloaded file. The extension will be '.xls'
            This parameter is inserted directly to the response's
            Content-Disposition, and must be escaped accordingly.

        sheetname: the name of the spreadsheet.

        table: a 2-dimensional iterable, in row-major order.

        header: an optional 2-dimensional iterable, in row-major order.

        include_totals: an optional boolean to include total values.

        total_label: Name of the total column, defaults to 'Total'

        grouper_col: Name of the group to subtotal values for (e.g. 'Site').

        value_col: Name of the column which holds the values to be summed
         (e.g.'Amount').

    Returns:

        A Microsoft Excel HttpResponse with appropriate content_type and
        Content-Disposition.
    """
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="%s.xls"' \
                                      % filename
    wb = xlwt.Workbook()
    ws = wb.add_sheet(sheetname)
    data_table = [list(x) for x in table]
    widths = dict()
    heights = dict()
    row_offset = 0
    max_column = 0

    if header is not None:
        for r, row in enumerate(header):
            for c, cell in enumerate(row):
                label = str(cell)
                height = get_height(label)
                width = get_width(label)
                ws.write(r, c, label)

                if height > heights.get(r, 0):
                    heights[r] = height
                    ws.row(r).height = height

                if width > widths.get(c, 0):
                    widths[c] = width
                    try:
                        ws.col(c).width = width
                    except ValueError:
                        ws.col(c).width = 8000

            row_offset += 1
        row_offset += 1

    first_row = row_offset
    r = 0
    for r, row in enumerate(data_table):
        r += row_offset
        for c, cell in enumerate(row):
            label = unicode(cell)
            height = get_height(label)
            width = get_width(label)
            date_style = xlwt.easyxf(num_format_str="YYYY-MM-DD")
            try:
                if is_datetime(cell):
                    try:
                        ws.write(r, c, cell, style=date_style)
                    except:
                        ws.write(r, c, label, style=date_style)
                else:
                        ws.write(r, c, cell)
            except:
                ws.write(r, c, label)

            if height > heights.get(r, 0):
                heights[r] = height
                ws.row(r).height = height

            if width > widths.get(c, 0):
                widths[c] = width
                try:
                    ws.col(c).width = width
                except ValueError:
                    ws.col(c).width = 8000

            if c > max_column:
                max_column = c
    row_offset = r + 1
    if include_totals:
        sub_total = 0
        sub_total_col = max_column + 1
        grand_total = 0

        # Locates the column indices from the title row of the data table.
        grouper_col = data_table[first_row].index(grouper_col.title())
        value_col = data_table[first_row].index(value_col.title())

        # Adds a new column, based on the provided title
        ws.write(first_row, sub_total_col, total_label)

        for r, row in list(enumerate(data_table)):
            if r >= 1:
                # We begin calculation at the first data row
                sub_total += data_table[r][value_col]
            if r >= 2:
                # If we're beyond the first data row and the grouper values
                # mismatch (indicating a new group) it returns the subtotal
                if data_table[r][grouper_col] != data_table[r - 1][grouper_col]:
                    ws.write(r - 1, sub_total_col, sub_total)
                    grand_total += sub_total
                    sub_total = 0
                else:
                    # Determines if we're at the end of the data table and
                    # inserts the totals
                    try:
                        data_table[r + 1][grouper_col]
                    except IndexError:
                        ws.write(r, sub_total_col, sub_total)
                        grand_total += sub_total
                        ws.write(r + 2, sub_total_col, grand_total)

        # Determines greatest text width between the title and the grand total
        # Grand total should be the widest tetx value in that column most of
        # the time
        width = max(get_width(total_label), get_width(str(grand_total)))
        ws.col(sub_total_col).width = width
    if footer is not None:
        for r, row in enumerate(footer):
            r += row_offset
            for c, cell in enumerate(row):
                label = str(cell)
                height = get_height(label)
                width = get_width(label)
                ws.write(r, c, label)

                if height > heights.get(r, 0):
                    heights[r] = height
                    ws.row(r).height = height

                if width > widths.get(c, 0):
                    widths[c] = width
                    try:
                        ws.col(c).width = width
                    except ValueError:
                        ws.col(c).width = 8000

                if c > max_column:
                    max_column = c
    wb.save(response)
    return response


def xls_multiple_worksheets_response(filename, data, padding=0):
    """
    Take a filename and a dictionary (data) and return a .xls response that
    can have multiple sheets.

    The user may indicate a style for a cell by passing in a dictionary with
    keys 'label' and 'style' instead of just a string.

    The user may provide a header for each sheet. A header is a set of cells
    under the key "header" that appears first in the sheet.

    data dict format:

    .. code-block:: python
        :linenos:

        mystyle = 'font: bold 1'
        data = {
            sheet_name: {
                'header': [
                    ['cell 1:1', 'cell 1:2'],
                    ['cell 2:1', {'label': 'cell 2:2', 'style': mystyle}]
                ],
                'table': [
                    [{'label': 'cell 3:1', 'style': mystyle}, 'cell 3:2'],
                    ['cell 4:1', 'cell 4:2']
                ]
            },
        }

    Styles are XFStyle strings. Comprehensive documentation for the format of
    these strings is difficult to find.

    Brief example:
        http://xlwt.readthedocs.org/en/latest/api.html#xlwt.Style.easyxf

    Our example:
        .. code-block:: python
            :linenos:

            style = 'font: bold 1,
                    'name Tahoma, ' \\
                    'height 160;' \\
                    'borders: left thick, right thick, top thick, ' \\
                    'bottom thick;  ' \\
                    'pattern: pattern solid, pattern_fore_colour  ' \\
                    'yellow,pattern_back_colour yellow'
    """
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename="%s.xls"' \
                                      % filename
    wb = xlwt.Workbook(style_compression=2)
    if not data:
        wb.add_sheet("Empty report")
        wb.save(response)
        return response
    for sheetname, content in data.items():
        table = content['table']
        ws = wb.add_sheet(sheetname.title(), cell_overwrite_ok=True)
        widths = dict()
        heights = dict()
        row_offset = 0
        max_column = 0

        if 'header' in data[sheetname] and data[sheetname]['header'] is not None:
            header = data[sheetname].pop('header')
            for r, row in enumerate(header):
                for c, cell in enumerate(row):
                    width, height = write_to_worksheet(ws, r, c, cell)
                    width += padding

                    if height > heights.get(r, 0):
                        heights[r] = height
                        ws.row(r).height = height

                    if width > widths.get(c, 0):
                        widths[c] = width
                        try:
                            ws.col(c).width = width
                        except ValueError:
                            ws.col(c).width = 8000

                row_offset += 1
            row_offset += 1

        data_table = [list(x) for x in table]
        for r, row in enumerate(data_table):
            r += row_offset
            for c, cell in enumerate(row):
                width, height = write_to_worksheet(ws, r, c, cell)
                width += padding

                if height > heights.get(r, 0):
                    heights[r] = height
                    ws.row(r).height = height

                if width > widths.get(c, 0):
                    widths[c] = width
                    try:
                        ws.col(c).width = width
                    except ValueError:
                        ws.col(c).width = 8000

                if c > max_column:
                    max_column = c
    wb.save(response)
    return response


def xlsx_response(filename, table, max_width=118, max_height=90):
    """Return a Microsoft Excel 2007+ file of the given table as an
     HttpResponse.

    Args:

        filename: the name of the downloaded file. The extension will be
        '.xlsx'. This parameter is inserted directly to the response's
        Content-Disposition, and must be escaped accordingly.

        table: a 2-dimensional iterable, in row-major order.

    Returns:

        A Microsoft Excel 2007+ HttpResponse with appropriate content_type and
        Content-Disposition.

    """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="%s.xlsx"' % filename
    wb = openpyxl.Workbook()
    ws = wb.active
    widths = dict()
    heights = dict()

    for r, row in enumerate(table, start=1):
        for c, cell in enumerate(row, start=1):
            ws_cell = ws.cell(row=r, column=c)
            ws_cell.value = cell
            if type(cell) in [str, unicode]:
                cell_str = ws_cell.value.encode('utf-8')
            elif type(cell) in [float]:
                ws_cell.number_format = '0.00'
                cell_str = str(ws_cell.value)
            else:
                cell_str = str(cell)

            widths[c] = min(max((widths.get(c, 0), len(cell_str))), max_width)
            cell_height = int(len(cell_str.split('\n')) * 15)
            heights[r] = min(max((heights.get(r, 0), cell_height)), max_height)

    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = width + 1

    for row, height in heights.items():
        ws.row_dimensions[row].height = height

    # Save to temporary file
    if settings.FILE_UPLOAD_TEMP_DIR:
        my_temp_file = tempfile.NamedTemporaryFile(
            suffix='.xlsx', dir=settings.FILE_UPLOAD_TEMP_DIR)
    else:
        my_temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx')
    print my_temp_file.name
    wb.save(my_temp_file.name)
    my_file = my_temp_file.file
    response.write(my_file.read())
    my_file.close()
    return response


def xlsx_multiple_worksheets_response(filename, data, max_width=118, max_height=90):
    """
        Takes a filename and an ordered dictionary (data) and returns an .xlsx response that
        can have multiple worksheets.

        data dict format:

        .. code-block:: python
            :linenos:

            data = {
                sheet_name1: {
                    'table': [
                        ['cell 1:1', 'cell 1:2', 'cell 1:3'],  # This is often the header row
                        ['cell 2:1', 'cell 2:2', 'cell 2:3'],
                        ['cell 3:1', 'cell 3:2', 'cell 3:3'],
                    ]
                },
                sheet_name2: {
                    'table': [
                        ['cell 1:1', 'cell 1:2'],
                        ['cell 2:1', 'cell 2:2'],
                        ['cell 3:1', 'cell 3:2'],
                        ['cell 4:1', 'cell 4:2'],
                    ]
                },
            }
        """
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="%s.xlsx"' % filename
    wb = openpyxl.Workbook()

    if not data:
        ws = wb.active
        ws.title("Empty Report")
    else:
        counter = 1
        for sheet_name, content in data.items():
            widths = dict()
            heights = dict()
            table = content['table']

            if counter == 1:
                ws = wb.active
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)

            counter += 1

            for r, row in enumerate(table, start=1):
                for c, cell in enumerate(row, start=1):
                    ws_cell = ws.cell(row=r, column=c)
                    ws_cell.value = cell
                    if type(cell) in [str, unicode]:
                        cell_str = ws_cell.value.encode('utf-8')
                    elif type(cell) in [float]:
                        ws_cell.number_format = '0.00'
                        cell_str = str(ws_cell.value)
                    else:
                        cell_str = str(cell)

                    widths[c] = min(max((widths.get(c, 0), len(cell_str))), max_width)
                    cell_height = int(len(cell_str.split('\n')) * 15)
                    heights[r] = min(max((heights.get(r, 0), cell_height)), max_height)

            for column, width in widths.items():
                ws.column_dimensions[get_column_letter(column)].width = width + 1

            for row, height in heights.items():
                ws.row_dimensions[row].height = height

    if settings.FILE_UPLOAD_TEMP_DIR:
        my_temp_file = tempfile.NamedTemporaryFile(
            suffix='.xlsx', dir=settings.FILE_UPLOAD_TEMP_DIR)
    else:
        my_temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx')
    print my_temp_file.name
    wb.save(my_temp_file.name)
    my_file = my_temp_file.file
    response.write(my_file.read())
    my_file.close()
    return response


def getattr_chain(obj, name_chain, suppress_attr_errors=False, sep='__'):
    """Apply getattr successively to a chain of attribute names.

    Argument 'name_chain' is a string containing sequence of attribute names
    to look up, starting with the initial object 'obj' and progressing through
    the chain. By default, a double underscore ('__') is expected to separate
    attribute names (as in Django's admin config and queryset keyword args),but
    any string may be specified in argument 'sep'. If 'sep' is None, argument
    'name_chain' is instead expected to be an already-separated iterable of
    attribute names.

    When evaluating a chain of attributes such as 'foo__bar__baz',in some cases
    'bar' may sometimes be None, such as in database models with nullable
    foreign keys. In order to simplify the process of attempting to look up
    such values, argument 'suppress_attr_errors' may be given: if it is True,
    any AttributeErrors raised by lookups on None (e.g., 'None.baz') will be
    caught, and the value None will be returned instead. (Attempted lookups of
    invalid names will still raise errors as usual.)
    Be aware, though, that specifying this option will result in the same
    behavior whether 'bar' or 'baz' is None.

    Note that while Django's uses of such string-specified attribute lookups
    are limited to database relations, this function performs just as well with
    regular object attributes, and even with properties.

    If a more complex lookup involving function calls or other logic is desired
    consider a lambda function, such as `lambda obj: obj.foo.bar.baz.qux()`.

    Args:

        obj: the object start the attribute lookup from.

        name_chain: a string containing a sequence of attribute names,separated
            by the value of argument 'sep'. May instead be an iterable of
            attribute names if 'sep' is None.

        suppress_attr_errors: if True, catches AttributeErrors raised from an
            attempted lookup on a None value anywhere in the attribute chain,
            and returns None instead of raising the exception.

        sep: the delimiting characters between the consecutive attribute names
            in argument 'name_chain'. Default is '__', but may be any string.
            If None, 'name_chain' is expected to be an iterable sequence of
            names, rather than a single string.

    Returns:
        The evaluation of the consecutive lookup of attributes in 'name_chain'.

    Example usage:
    ::

        >>> class Obj(object): pass
        >>> obj, obj.foo = Obj(), Obj()

        >>> obj.foo.bar = None
        >>> getattr_chain(obj, 'foo__bar')
        >>> # None returned.
        >>> getattr_chain(obj, 'foo__bar__baz')
        Traceback (most recent call last):
            ...
        AttributeError: 'NoneType' object has no attribute 'baz'
        >>> getattr_chain(obj, 'foo__bar__baz', suppress_attr_errors=True)
        >>> # None returned; no exception raised.

        >>> obj.foo.bar = 'spam'
        >>> getattr_chain(obj, 'foo__bar')
        'spam'
        >>> getattr_chain(obj, 'foo__bar__baz')
        Traceback (most recent call last):
            ...
        AttributeError: 'str' object has no attribute 'baz'
        >>> getattr_chain(obj, 'foo__bar__baz', suppress_attr_errors=True)
        Traceback (most recent call last):
            ...
        AttributeError: 'str' object has no attribute 'baz'
        >>> # Only AttributeErrors from NoneType are suppressed.
    """
    names = name_chain if sep is None else name_chain.split(sep)

    for name in names:
        try:
            obj = getattr(obj, name)
        # TODO: consider catching ObjectDoesNotExist as well.
        except AttributeError:
            if suppress_attr_errors and obj is None:
                return None
            else:
                # If suppress_attr_errors is not set, or if the error
                # was due to an invalid field name (rather than a valid
                # field having the value None), re-raise this exception.
                raise
    return obj


def generate_basic_table(columns, data):
    """Generate a table of functions applied to data objects.

    Argument 'columns' is an iterable of 2-tuples of the form (title, function)
    where 'title' is the column title and 'function' is a single-parameter
    function that will be applied to each data element to create the column.

    Returns a 2-dimensional row-major-ordered generator. The first row is the
    column titles; subsequent rows are evaluated functions of the data points.

    Args:
        columns: an iterable of pairs (title, function):
            title: the string to appear at the top of the column.
            function: a callable to be applied to each datum in the column.
        data: a QuerySet, list, or other iterable of arbitrary objects that can
            be passed to the provided functions.
    Yields:
        rows of the table:
            first row: the given column titles.
            subsequent rows: the given column functions applied to each datum.

    Example usage:
        >>> columns = [('Numbers', lambda x: x), ('Squares', lambda x: x ** 2)]
        >>> data = [1, 2, 3, 4, 5]
        >>> list(list(row) for row in generate_basic_table(columns, data))
        [['Numbers', 'Squares'], [1, 1], [2, 4], [3, 9], [4, 16], [5, 25]]

    Author:
        Fredrick Wagner
    """
    titles, functions = [], []
    columns = list(columns)
    if columns:
        titles, functions = zip(*columns)
    yield titles
    for obj in data:
        yield (f(obj) for f in functions)  # One row of the table.


def generate_table(columns, data, model=None, capitalize_title=True,
                   remove_nones=False, **kwargs):
    """Wrapper around generate_basic_table with fancier column specifications.

    Argument 'columns' is an iterable of specifications for table columns.
    Each specification is in one of three forms:

        - a tuple (title, func), where 'title' is the column title and 'func'
            is a function that will be applied to each datum in the column.
            Specifications in this format will be passed through unchanged.

        - a tuple (title, attr). The column function will be a lookup of the
            provided attribute, which may be a model field, an ordinary object
            attribute, or a property.

        - a string, which is an attribute name as specified above. If the
            attribute is a model field and the 'model' arg is given, this
            column's title will be the model's verbose_name; otherwise,
            the title will be the string, with single and double underscores
            converted to single spaces.

    See also getattr_chain, which this function uses to look up attributes
    given by the latter two types of column specifications. Any keyword args
    are passed through to it.

    Args:

        columns: an iterable of column specifications.

        model: the model to look up field names from. Only used to get verbose
            field names for column titles, and not used at all if all column
            specifications include titles.

        capitalize_title: if True (the default), derived column names will have
            their first letter capitalized.

        remove_nones: if True, 'None' results from column functions will be
            replaced with the empty string.

        kwargs: passed through to function 'getattr_chain'.

    Yields:
        Column specifications in the form (title, func).

    Example:
    ::

        def scholarship_report_view(request):
            table = generate_table([
                'id',
                'parent',
                ('Submission Date', 'submission_date'),
                ('Email Address', Scholarship.get_user_email),
                ('Random Numbers', lambda _: random.random()),
            ], data=Scholarship.objects.all(), model=Scholarship)
            return csv_response('Scholarship Information', table)


    Author:

        Fredrick Wagner

    """

    def filter_nones(func):
        def wrapper(*args, **kwargs):
            result = func(*args, **kwargs)
            return result if result is not None else ''

        return wrapper

    def converted_columns():
        """Convert column specifications to an identical format."""
        for column in columns:
            # First, figure out whether this column specification is a single
            # string or a tuple (title, spec).
            if type(column) == str:
                if column == '__str__':
                    title = model._meta.verbose_name
                else:
                    # This column is an attribute name.
                    try:
                        # If this attribute is a model field,
                        # retrieve its verbose_name from the model.
                        title = model._meta.get_field(column).verbose_name
                    except (AttributeError, FieldDoesNotExist):
                        # If no model was specified or no such field was found,
                        # fall back on the attribute name, with underscores
                        # (single or double) replaced with (single) spaces.
                        title = column.replace('__', ' ').replace('_', ' ')
                if capitalize_title:
                    title = capfirst(title)
                spec = column
            else:
                title, spec = column
            # Now figure out what to do with the spec.
            if type(spec) == str:
                if spec == '__str__':
                    func = str
                else:
                    # Assume this spec is an attribute name.
                    # Replace it with a function to get that attribute.
                    func = partial(getattr_chain, name_chain=spec, **kwargs)
            else:
                # Assume this spec is a ready-to-go function.
                func = spec
            if remove_nones:
                func = filter_nones(func)
            yield (title, func)

    return generate_basic_table(converted_columns(), data)
